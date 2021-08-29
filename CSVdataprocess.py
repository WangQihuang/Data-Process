# 作者：王启煌
# 创建时间：2021/08/28
# 程序简介：处理ICPMS8900质谱仪数据
#######################################################################################################

# 导入模块
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill
import xlrd
import pandas as pd

# 原始文件路径
# inputfilepass = r'C:\Users\Administrator\Desktop\data.csv'
# outputfilepass = r'C:\Users\Administrator\Desktop\ICPMS_data.xlsx'

#######################################################################################################

class dataprocess():
    def csvprocess(self,stdcalnum,plnum,inputfilepass,outputfilepass):
        # 将csv文件转为xisx文件
        csv = pd.read_csv(inputfilepass, encoding='utf-8')
        csv.to_excel(outputfilepass, sheet_name='ICPMS oringin data')

        # 读取原始数据表
        data = xlrd.open_workbook(outputfilepass)
        table = data.sheet_by_name('ICPMS oringin data')

        # 获取文件行列数
        wb = load_workbook(outputfilepass)
        ws = wb.active
        row_max = ws.max_row
        col_max = ws.max_column
        # print(row_max)
        # print(col_max)
        # count = row_max

        # 计算有效样品个数
        samplenum = int((row_max - stdcalnum - 2) / plnum)

        # 计算元素个数
        elementnum = int((col_max - 8) / 2)
        # print(elementnum)

        #######################################################################################################

        # 新建Step1工作表，用于存放筛选的有效列
        wb = openpyxl.load_workbook(outputfilepass)
        wb.create_sheet(title='Step1', index=0)

        # 复制文件名列到Step1工作表
        ws = wb['ICPMS oringin data']
        sheet = wb.active
        copycol = []
        i = 0
        while i < row_max:
            i = i + 1
            copycol.insert(i, ws.cell(i, 8).value)
        ws = wb['Step1']
        sheet = wb.active
        i = 0
        while i < row_max:
            ws.cell(i + 1, 1).value = copycol[i]
            i = i + 1

        # 复制元素浓度数据列到Step1工作表
        copycolnum = 9
        pastecolnum = 2
        while copycolnum < col_max:
            ws = wb['ICPMS oringin data']
            sheet = wb.active
            copycol = []
            i = 0
            while i < row_max:
                i = i + 1
                copycol.insert(i, ws.cell(i, copycolnum).value)
            ws = wb['Step1']
            sheet = wb.active
            i = 0
            while i < row_max:
                if copycol[i] == '<0.000':
                    copycol[i] = '0.000'
                    ws.cell(i + 1, pastecolnum).value = copycol[i]
                elif copycol[i] == None:
                    copycol[i] = '0.000'
                    ws.cell(i + 1, pastecolnum).value = copycol[i]
                else:
                    ws.cell(i + 1, pastecolnum).value = copycol[i]
                i = i + 1
            copycolnum = copycolnum + 2
            pastecolnum = pastecolnum + 1
        wb.save(outputfilepass)

        #######################################################################################################

        # 新建Step2工作表，用于存放筛选的有效行
        wb = openpyxl.load_workbook(outputfilepass)
        wb.create_sheet(title='Step2', index=0)

        # 复制表头到Step2工作表
        ws = wb['Step1']
        sheet = wb.active
        copyrow = []
        i = 0
        while i < elementnum + 1:
            i = i + 1
            copyrow.insert(i, ws.cell(1, i).value)
        ws = wb['Step2']
        sheet = wb.active
        i = 0
        while i < elementnum + 1:
            ws.cell(1, i + 1).value = copyrow[i]
            i = i + 1

        # 复制Step1工作表中的有效行到Step2工作表
        copyrownum = stdcalnum + 3
        pasterownum = 2
        while copyrownum < row_max + 1:
            ws = wb['Step1']
            sheet = wb.active
            copyrow = []
            i = 0
            while i < row_max + 1:
                i = i + 1
                copyrow.insert(i, ws.cell(copyrownum, i).value)
            # print(copyrow)
            ws = wb['Step2']
            sheet = wb.active
            i = 0
            while i < row_max + 1:
                ws.cell(pasterownum, i + 1).value = copyrow[i]
                i = i + 1
            copyrownum = copyrownum + 1
            pasterownum = pasterownum + 1
        wb.save(outputfilepass)

        #######################################################################################################
        # 设置单元格格式
        Color = ['ffffff', '000000', 'ff0000', 'ffff00', '87cefa', '90ee90', '', '']
        # 颜色注释[   白      黑      红       黄     淡蓝色   淡绿色]
        # 字体类型
        font1 = Font(u'Times New Roman', size=11, bold=False, italic=False, strike=False, color=Color[1])
        # 填充类型
        fill1 = PatternFill('solid', fgColor=Color[3])
        fill2 = PatternFill('solid', fgColor=Color[4])
        fill3 = PatternFill('solid', fgColor=Color[5])

        # 新建Data Process工作表，用于存放最终处理的数据
        wb = openpyxl.load_workbook(outputfilepass)
        wb.create_sheet(title='Data Process', index=0)

        # 在Data Process工作表中创建样品名称列（有效样品，已经计算了平均值和标准差）
        ws = wb['Data Process']
        sheet = wb.active
        ws.cell(1, 1).value = 'Sample Name'
        ws.cell(1, 1).font = font1
        ws.cell(1, 1).fill = fill1
        ws.cell(2, 1).value = 'Data Type'
        ws.cell(2, 1).font = font1
        ws.cell(2, 1).fill = fill1
        i = 1
        while i < samplenum + 1:
            ws.cell(i + 2, 1).value = 'Test Sample ' + str(i)
            ws.cell(i + 2, 1).font = font1
            ws.cell(i + 2, 1).fill = fill1
            i = i + 1
        sheet.column_dimensions['A'].width = 20

        wb.save(outputfilepass)

        #######################################################################################################

        # 载入样品单元素均值，标准差计算模块  循环遍历所有元素
        # copycolnum为所需复制的Step2工作表中的列号  每循环一次+1
        copycolnum = 2
        while copycolnum < elementnum + 2:
            # while copycolnum < 3:
            # 新建Step3工作表，用于计算平行样品数据
            wb = openpyxl.load_workbook(outputfilepass)
            wb.create_sheet(title='Step3', index=0)

            # 复制表头
            pastecolnum = 1
            ws = wb['Step2']
            sheet = wb.active
            copycol = []
            copycol.insert(1, ws.cell(1, copycolnum).value)
            while pastecolnum < plnum + 3:
                ws = wb['Step3']
                sheet = wb.active
                ws.cell(1, pastecolnum).value = copycol[0]
                pastecolnum = pastecolnum + 1
            wb.save(outputfilepass)

            # 添加注释（第2行）
            annotation = []
            i = 1
            while i < plnum + 1:
                annotation.append('Parallel Sample ' + str(i))
                i = i + 1
            annotation.append('Average')
            annotation.append('Stdev')

            i = 0
            while i < plnum + 2:
                ws = wb['Step3']
                sheet = wb.active
                ws.cell(2, i + 1).value = annotation[i]
                i = i + 1
            wb.save(outputfilepass)

            # 错位复制数据
            j = 1
            while j < plnum + 1:
                pastecolnum = 1
                ws = wb['Step2']
                sheet = wb.active
                copycol = []
                i = 1
                while i < row_max:
                    copycol.insert(i, ws.cell(i, copycolnum).value)
                    i = i + 1
                # print(copycol)
                i = 0
                while i < row_max - stdcalnum - 2:
                    ws = wb['Step3']
                    sheet = wb.active
                    ws.cell(i + 3, j).value = copycol[i + j]
                    i = i + 1
                j = j + 1
            wb.save(outputfilepass)

            # 计算平均值
            row = 3
            while row < row_max - stdcalnum - plnum + 2:
                i = 0
                ws = wb['Step3']
                sheet = wb.active
                copyrow = []
                while i < plnum:
                    copyrow.insert(i, float(ws.cell(row, i + 1).value))
                    i = i + 1
                # print(copyrow)
                # 求平均值
                copyrow_mean = np.mean(copyrow)
                # 求标准差
                copyrow_std = np.std(copyrow, ddof=1)
                # print("平均值为：%f" % copyrow_mean)
                # print("标准差为:%f" % copyrow_std)
                ws.cell(row, plnum + 1).value = copyrow_mean
                ws.cell(row, plnum + 2).value = copyrow_std
                row = row + 1
            wb.save(outputfilepass)

            # 将Step3计算完的结果复制到Data Process工作表中
            # 复制表头
            ws = wb['Step3']
            sheet = wb.active
            element = ws.cell(1, 1).value
            copyrow = []
            i = 0
            while i < 2:
                copycol.insert(i, ws.cell(i + 1, plnum + 1).value)
                i = i + 1
            ws = wb['Data Process']
            sheet = wb.active
            i = 0
            while i < 2:
                ws.cell(i + 1, copycolnum * 2 - 2).value = copycol[i]
                ws.cell(i + 1, copycolnum * 2 - 2).font = font1
                ws.cell(i + 1, copycolnum * 2 - 2).fill = fill1
                i = i + 1
            ws = wb['Step3']
            sheet = wb.active
            copyrow = []
            i = 0
            while i < 2:
                copycol.insert(i, ws.cell(i + 1, plnum + 2).value)
                i = i + 1
            ws = wb['Data Process']
            sheet = wb.active
            i = 0
            while i < 2:
                ws.cell(i + 1, copycolnum * 2 - 1).value = copycol[i]
                ws.cell(i + 1, copycolnum * 2 - 1).font = font1
                ws.cell(i + 1, copycolnum * 2 - 1).fill = fill1
                i = i + 1
            # 复制平均值列
            ws = wb['Step3']
            sheet = wb.active
            copyrow = []
            i = 0
            while i < samplenum:
                copycol.insert(i, ws.cell(plnum * i + 3, plnum + 1).value)
                i = i + 1
            ws = wb['Data Process']
            sheet = wb.active
            i = 0
            while i < samplenum:
                ws.cell(i + 3, copycolnum * 2 - 2).value = copycol[i]
                ws.cell(i + 3, copycolnum * 2 - 2).font = font1
                ws.cell(i + 3, copycolnum * 2 - 2).fill = fill2
                i = i + 1

            # 复制标准差列
            ws = wb['Step3']
            sheet = wb.active
            copyrow = []
            i = 0
            while i < samplenum:
                copycol.insert(i, ws.cell(plnum * i + 3, plnum + 2).value)
                i = i + 1
            ws = wb['Data Process']
            sheet = wb.active
            i = 0
            while i < samplenum:
                ws.cell(i + 3, copycolnum * 2 - 1).value = copycol[i]
                ws.cell(i + 3, copycolnum * 2 - 1).font = font1
                ws.cell(i + 3, copycolnum * 2 - 1).fill = fill3
                i = i + 1
            wb.save(outputfilepass)

            # 删除工作表Step3
            wb = openpyxl.load_workbook(outputfilepass)
            ws = wb['Step3']
            wb.remove(ws)
            wb.save(outputfilepass)
            print('Element "' + element + '" Cauculate Successfully!')

            # 下一个循环
            copycolnum = copycolnum + 1

        # 删除工作表Step1,Step2
        wb = openpyxl.load_workbook(outputfilepass)
        ws = wb['Step1']
        wb.remove(ws)
        wb.save(outputfilepass)
        # print('Delte Step1 Successfully!')
        wb = openpyxl.load_workbook(outputfilepass)
        ws = wb['Step2']
        wb.remove(ws)
        wb.save(outputfilepass)
        # print('Delte Step2 Successfully!')

        # print(table.row(1))
        wb.save(outputfilepass)



if __name__ == '__main__':
    username = input('输入用户名:')
    # 输入标准曲线样品个数
    stdcalnum = int(input('输入标准曲线样品个数:'))
    # 输入平行样品个数
    plnum = int(input('输入平行样品个数:'))
    # 输入要处理的文件路径
    inputfilepass = input('输入要处理的文件路径:')
    # 输入处理后存放文件的路径
    outputfilepass = r'C:\Users\\'+ username + r'\Desktop\ICPMS_data.xlsx'
    dataprocess.csvprocess('',stdcalnum,plnum,inputfilepass,outputfilepass)